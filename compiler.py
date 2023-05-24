import re
import os
import openpyxl

#Memory which stores sheet data
mem = {}

def read_excel_data(filename, start_row=1, end_row=None, start_column=1, end_column=None):
    # Load the workbook
    workbook = openpyxl.load_workbook(filename)

    # Select the sheet you want to read from
    sheet = workbook.active

    # Determine the maximum number of rows and columns if not provided
    if start_row is None or end_row is None:
        start_row = 1
        end_row = sheet.max_row
    if start_column is None or end_column is None:
        start_column = 1
        end_column = sheet.max_column
    
    #print(start_row, " ", end_row, " ", start_column, " ", end_column)

    arr = []

    # Iterate over the specified range of cells and print the cell value
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_column, max_col=end_column, values_only=True):
        r = []
        for cell in row:
            r.append(cell)
        arr.append(r)
    mem[filename] = arr

def interpreter(file_contents):
    # Split the text into an array of lines
    lines = file_contents.split("\n")

    #Remove empty lines
    lines = [element for element in lines if element != ""]
    # Iterate through the lines
    for line in lines:
        if line.strip()[0] == "#":
            continue
        #print(line)
        
        #Check for read
        read_pattern = r"read\s+(\w+\.\w+)\s+(?:(?:rows|cols)\s+(\d+)-(\d+))(?:\s*(?:rows|cols)\s+(\d+)-(\d+))?"
        read_matches = re.match(read_pattern, line)

        if read_matches:
            filename = read_matches.group(1)
            start_row = int(read_matches.group(2)) if read_matches.group(2) else None
            end_row = int(read_matches.group(3)) if read_matches.group(3) else None
            start_column = int(read_matches.group(4)) if read_matches.group(4) else None
            end_column = int(read_matches.group(5)) if read_matches.group(5) else None

            read_excel_data(filename, start_row, end_row, start_column, end_column)
        
        #Check for output
        output_pattern = r"output\s+memory(?:\s+(\w+(\.\w+)?))?"
        output_matches = re.match(output_pattern, line)
        
        if output_matches:
            if output_matches.group(1):
                print(mem[output_matches.group(1)])
            else:
                print(mem)

                
# Main program
def main():
    file_path = "sample.dmaa"
    #file_path = input('Enter the file path: ')

    # Check file extension
    file_name, file_extension = os.path.splitext(file_path)
    if file_extension != '.dmaa':
        print('Invalid file extension. Please provide a .dmaa file.')
        return

    try:
        with open(file_path, 'r') as file:
            file_contents = file.read()
    except FileNotFoundError:
        print('File not found.')
        return

    interpreter(file_contents)

if __name__ == '__main__':
    main()